import os
import sys
import shutil
import openpyxl


mode = sys.argv[1]
dir_name = sys.argv[2]
num_p = int(sys.argv[3])
num_w = int(sys.argv[4])

dir_path = "experiment/"+dir_name
new_excel_path = dir_path+"/input/"+dir_name+".xlsx"
if mode=="create":    
    # try:
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
        os.makedirs(dir_path+"/input")
        os.makedirs(dir_path+"/output")
        os.makedirs(dir_path+"/result")
        print('input 폴더에 input.xlsx을 넣어야 함')
        shutil.copy("standard.xlsx", dir_path+"/input/"+dir_name+".xlsx")
        wb = openpyxl.load_workbook(new_excel_path)
        labor_sheet = wb['labor_raw']
        labor_sheet.cell(row=2, column=3, value=num_p)
        labor_sheet.cell(row=3, column=3, value=num_w)
        wb.save(dir_path+"/input/"+dir_name+".xlsx")
        wb.close()
    else:
        print('already exist')
    # except OSError:
    #     print ('Error: Creating directory. ' +  dir_path)

    
elif mode=="delete":
    shutil.rmtree(dir_path)