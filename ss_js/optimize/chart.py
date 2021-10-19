import plotly as py
import plotly.figure_factory as ff
import plotly.express as px
import pandas as pd
from ss_js import utils
from ss_js.schedule.schedule import Schedule
import json
import openpyxl
import plotly.graph_objects as go
import math


def task_list_to_excel(data_path, result_path):
    with open(data_path, 'r', encoding="utf-8") as f:
        task_list_data = json.load(f)
    col_names = ["task_id", "workpackage_id", "section", "space_id", "start_value", "end_value","duration", "labor_type", "num_labor", "productivity", "quantity"]
    wb = openpyxl.Workbook()
    wb.create_sheet('task_list')
    wb.create_sheet('task_list(day)')
    task_list_sheet = wb['task_list']
    task_list_day_sheet = wb['task_list(day)']   

    for seq, name in enumerate(col_names):
        task_list_sheet.cell(row=1, column=seq+1, value=name)
        task_list_day_sheet.cell(row=1, column=seq+1, value=name)
    
    for idx, data in enumerate(task_list_data):
        for seq, name in enumerate(col_names):            
            task_list_sheet.cell(row=idx+2, column=seq+1, value=data[name])
            if name in ["start_value", "end_value","duration"]:
                day_data = math.ceil(data[name]/480)
                task_list_day_sheet.cell(row=idx+2, column=seq+1, value=day_data)
            else:
                task_list_day_sheet.cell(row=idx+2, column=seq+1, value=data[name])
    
    wb.save(result_path+".xlsx")
    wb.close()
    return


def get_labor_time_col_list(labor_data):
    labor_list = ["time"]
    for labor_info in labor_data.values():
        for labor_type in labor_info.keys():
            if labor_type not in labor_list:
                labor_list.append(labor_type)
    return labor_list

def labor_time_to_json(data_path, result_path):
    with open(data_path, 'r', encoding="utf-8") as f:
        labor_time_data = json.load(f)
    
    new_json_data = {}
    labor_time_col_list = get_labor_time_col_list(labor_time_data)    
    
    for seq, name in enumerate(labor_time_col_list):
        new_json_data[name] = []
        
    idx = 0    
    for time, labor_info in labor_time_data.items():
        new_json_data["time"].append(int(time))
        for seq, labor_type_id in enumerate(labor_time_col_list):
            if seq == 0:
                pass
            else: 
                if labor_type_id in labor_info.keys():
                    num_labor = labor_info[labor_type_id]                    
                    new_json_data[labor_type_id].append(num_labor)
                else:                    
                    new_json_data[labor_type_id].append(0)
        idx += 1    
    with open(result_path+".json", 'w', encoding="utf-8") as f:
         json.dump(new_json_data, f, ensure_ascii=False)
    return

def labor_time_to_excel(data_path, result_path):    
    labor_time_to_json(data_path, result_path)
    with open(result_path+".json", 'r', encoding="utf-8") as f:
        labor_time_data = json.load(f)

    wb = openpyxl.Workbook()    
    wb.create_sheet('labor_time')        
    wb.create_sheet('labor_time_day')   
    labor_time_sheet = wb['labor_time']
    labor_time_day_sheet = wb['labor_time_day']

    for idx, col_name in enumerate(labor_time_data.keys()):
        labor_time_sheet.cell(row=1, column=idx+1, value=col_name)
        labor_time_day_sheet.cell(row=1, column=idx+1, value=col_name)
        
    col_idx = 1    
    for col_name, row_list in labor_time_data.items():
        day_data = []        
        for row_idx, row_value in enumerate(row_list):            
            labor_time_sheet.cell(row=row_idx+2, column=col_idx, value=row_value)            
            day_data.append(row_value)
            if col_name == "time":
                if row_idx % 480 == 479:
                   labor_time_day_sheet.cell(row=(row_idx//480)+2, column=1, value=row_idx//480) 
            else:                
                if row_idx % 480 == 479:
                    day = row_idx / 480
                    day += 1
                    num_labor = max(day_data)
                    labor_time_day_sheet.cell(row=(row_idx//480)+2, column=col_idx, value=num_labor)
                    day_data.clear()
        col_idx += 1

    wb.save(result_path+".xlsx")
    wb.close()    
    return


def create_gantt(data_path, out_dir, schedule: Schedule):
    with open(data_path, 'r', encoding="utf-8") as f:
        json_data = json.load(f)

    figure_data = []
    workpackage_list = []
    section_list = []
    for data in json_data:        
        section = data['section']
        workpackage = data['workpackage_id']
        figure_data.append(dict(
            Task=data["task_id"],
            Start=data["start_value"],
            Finish=data['end_value'],
            Workpackage=data['workpackage_id'],
            Section=data['section']
        ))
        if section not in section_list:
            section_list.append(section)
        if workpackage not in workpackage_list:
            workpackage_list.append(workpackage)

    df = pd.DataFrame(figure_data)
    df['delta'] = df['Finish'] - df['Start']
    pyplt = py.offline.plot
    wp_colors = {}
    # for workpackage_id in schedule.workpackage_list():
    for workpackage_id in workpackage_list:
        wp_colors[workpackage_id] = utils.random_rgb_txt()
    
    sc_colors = {}
    # for section in schedule.section_list():
    for section in section_list:
        sc_colors[section] = utils.random_rgb_txt()

    fig = ff.create_gantt(
        df,
        colors= wp_colors,        
        index_col='Workpackage',
        show_colorbar = True,
        group_tasks = True
    )
    fig.layout.xaxis.type = 'linear'

    fig_2 = px.timeline(df, x_start="Start", x_end="Finish", color="Section")
    fig_2.layout.xaxis.type ='linear'
    fig_3 = ff.create_gantt(
        df,
        colors = sc_colors,
        index_col='Section',
        show_colorbar=True,
        group_tasks=True
    )
    fig_3.layout.xaxis.type = 'linear'

    pyplt(fig, filename=out_dir+"workpackage_chart.html", auto_open=False)    
    pyplt(fig_3, filename=out_dir+"section_chart.html", auto_open=False)  
    return


def get_labor_list(json_data):
    labor_id_list = []
    for data in json_data.values():        
        for labor_id, num_labor in data.items():
            if labor_id not in labor_id_list:
                labor_id_list.append(labor_id)
    return labor_id_list
    

def create_new_labor_dict(json_data):    
    labor_id_list = get_labor_list(json_data)
    new_json_data = {}
    for labor_id in labor_id_list:
        new_json_data[labor_id] = []

    new_json_data["time"] = []

    t = 0
    for time, data in json_data.items():
        new_json_data["time"].append(time)
        for labor_id in labor_id_list:
            if labor_id == "time":
                continue
            if labor_id in data.keys():
                num_labor = data[labor_id]
                new_json_data[labor_id].append(num_labor)
            else:
                                    
                new_json_data[labor_id].append(0)            
        t += 1
    
    return new_json_data

def create_labor_chart(data_path, out_dir):
    with open(data_path, 'r', encoding="utf-8") as f:
        json_data = json.load(f)

    labor_id_list = get_labor_list(json_data)
    new_data = create_new_labor_dict(json_data)
    
    
    # new_data = pd.DataFrame(new_data)
    
   
    fig = go.Figure()
    for labor_id in labor_id_list:
        fig.add_trace(go.Scatter(x=new_data["time"], y=new_data[labor_id],
                            mode='lines', # Line plot만 그리기
                            name=labor_id))
        # fig.add_trace(go.Bar(
        #     x=new_data["time"],
        #     y=new_data[labor_id],
        #     name=labor_id
        # ))
    pyplt = py.offline.plot
    pyplt(fig, filename=out_dir+"labor_chart.html", auto_open=False)  

    return
        
