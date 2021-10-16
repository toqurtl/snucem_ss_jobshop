import collections
from ortools.sat.python import cp_model
import openpyxl
import json

dir_txt = "data/input.xlsx"
# read_space_raw
def get_section_list(dir_txt):
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel['space_raw']
    ctn = 0
    section_list = []
    for row in excel_ws.iter_rows():
        if ctn > 0:
            module_type, zone, num_section, quantity = \
                row[0].value, row[1].value, int(row[2].value), int(row[3].value)
            section_list.append({
                "section": zone,
                "module_type": module_type,
                "num_section": num_section,
                "quantity": quantity
            })
        ctn += 1

    return section_list
    
def get_workpackage_list(dir_txt):
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel['workpackage_raw']
    ctn = 0
    wp_list = []
    for row in excel_ws.iter_rows():
        if ctn > 0:
            wp_id, wp_name = row[0].value, row[1].value
            wp_list.append({
                "workpackage_id": wp_id,
                "workpacakage_name": wp_name
            })
        ctn += 1

    return wp_list


def get_space_list(dir_txt):
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel['space_raw']
    ctn = 0
    space_list = []
    for row in excel_ws.iter_rows():
        if ctn > 0:
            module_type, zone, num_section, quantity, fixed_start, fixed_finish = \
                row[0].value, row[1].value, int(row[2].value), int(row[3].value), int(row[4].value), int(row[5].value)
            
            space_list.append(
                {
                    "space_id": zone,
                    "module_type": "module",
                    "section": zone,
                    "detail_section": 0,
                    "is_module": True,
                    "quantity": 5,
                    "fixed_start": fixed_start,
                    "fixed_finish": fixed_finish
                }
            )
            for idx in range(1, num_section + 1):
                space_list.append(
                    {
                        "space_id": zone+"_"+str(idx),
                        "module_type": module_type,
                        "section": zone,
                        "detail_section": idx,
                        "is_module": False,
                        "quantity": round(quantity / num_section),
                        "fixed_start": 0,
                        "fixed_finish": 0
                    }
                )
        ctn += 1
    return space_list


def get_alt_id_list(task_id_arg, dir_txt):
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel['alt_raw']    
    alt_id_list = []
    for idx, row in enumerate(excel_ws.iter_rows()):
        if idx > 0:
            task_id, alt_id, required_labor, required_number, productivity, is_productivity, fixed_duration = \
                row[0].value, row[1].value, row[2].value, row[3].value, float(row[4].value), bool(row[5].value), int(row[6].value)
            if task_id == task_id_arg:
                alt_id_list.append({
                    "alt_id": alt_id,
                    "required":{
                        required_labor: required_number
                    },
                    "productivity": productivity,
                    "is_productivity": is_productivity,
                    "fixed_duration": fixed_duration
                })
    
    return alt_id_list


def get_tasktype_list(dir_txt):
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel['tasktype_raw']
    ctn = 0
    tasktype_list = []    
    for row in excel_ws.iter_rows():        
        if ctn > 0:
            work_id, workpackage_id, workpackage_name, detail_id, detail_name, is_module \
                = row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, bool(row[5].value)
            
            tasktype_list.append({
                "task_id": work_id,
                "workpackage_id": workpackage_id,
                "workpackage_name": workpackage_name,
                "detail_id": detail_id,
                "task_name": detail_name,
                "is_module": is_module,
                "labor_set": get_alt_id_list(work_id, dir_txt)
                
            })
        ctn += 1
    return tasktype_list

def get_low_space_list(space_id, space_list):
    low_space_list = []
    for space in space_list:
        if space["section"] == space_id and not space["is_module"]:
            low_space_list.append(space["space_id"])
    return low_space_list


def get_labor_type_list(dir_txt):
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel['labor_raw']    
    labor_list = []
    ctn = 0    
    for row in excel_ws.iter_rows():        
        if ctn > 0:
            labor_id, labor_name, number \
                = row[0].value, row[1].value, int(row[2].value)
            
            labor_list.append({
                "labor_id": labor_id,
                "labor_name": labor_name,
                "number": number
            })
        ctn += 1
    
    return labor_list


def get_work_list(dir_txt):
    space_list = get_space_list(dir_txt)
    tasktype_list = get_tasktype_list(dir_txt)
    work_list = []
    for space in space_list:    
        space_id = space["space_id"]    
        for tasktype in tasktype_list:
            if space["is_module"]:
                if tasktype["is_module"]:
                    work_id = space_id + "_" + tasktype["task_id"]                
                    work_list.append({
                        "work_id": work_id,
                        "task_type_id": tasktype["task_id"],
                        "quantity": space["quantity"],                    
                        "space_id": get_low_space_list(space_id, space_list),
                        "section": space["section"],
                        "workpackage_id": tasktype["workpackage_id"],
                        "is_module": tasktype["is_module"],
                        "fixed_start": space["fixed_start"],
                        "fixed_finish": space["fixed_finish"]
                    })
                    
            else:
                if not tasktype["is_module"]:
                    work_id = space_id + "_" + tasktype["task_id"]
                    work_list.append({
                        "work_id": work_id,
                        "task_type_id": tasktype["task_id"],
                        "quantity": space["quantity"],
                        "space_id": [space_id],
                        "section": space["section"],
                        "workpackage_id": tasktype["workpackage_id"],
                        "is_module": tasktype["is_module"]
                    })
    return work_list


def get_work_of_section(section, work_list):
    new_work_list = []
    for work in work_list:
        if work["section"] == section:
            new_work_list.append(work)
    return new_work_list


def get_work_of_space(space_id, work_list):
    new_work_list = []    
    for work in work_list:        
        if len(work["space_id"]) == 1 and work["space_id"][0] == space_id:
            new_work_list.append(work)
    return new_work_list


def get_work(task_type_id, work_list):
    new_work_list = []
    for work in work_list:        
        if task_type_id == work["task_type_id"]:
            new_work_list.append(work)
    return new_work_list


def get_dependency_list(dir_txt):    
    section_list = get_section_list(dir_txt)
    space_list = get_space_list(dir_txt)
    work_list = get_work_list(dir_txt)
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel['dep_raw']    
    dep_list = []
    # 모듈 내 세부공종들
    for space in space_list:
        if not space["is_module"]:
            works_of_space = get_work_of_space(space["space_id"], work_list)                         
            ctn = 0
            for row in excel_ws.iter_rows():        
                if ctn > 0:
                    pre, pre_workpackage, suc, suc_workpackage = \
                        row[0].value, row[1].value, row[2].value, row[3].value
                    if pre_workpackage != "Z":                    
                        pre_work = get_work(pre, works_of_space)[0]
                        suc_work = get_work(suc, works_of_space)[0]                  
                        dep_list.append([pre_work["work_id"], suc_work["work_id"]])
                ctn += 1

    
    for section in section_list:
        works_of_section = get_work_of_section(section["section"], work_list)
        ctn = 0
        for row in excel_ws.iter_rows():
            if ctn > 0:
                pre, pre_workpackage, suc, suc_workpackage = \
                    row[0].value, row[1].value, row[2].value, row[3].value
                # 모듈과 그 안의 공종들
                if pre_workpackage == "Z" and suc_workpackage != "Z":
                    
                    pre_work = get_work(pre, works_of_section)[0]
                    suc_work_list = get_work(suc, works_of_section)
                    for suc_work in suc_work_list:
                        dep_list.append([pre_work["work_id"], suc_work["work_id"]])
                        pass
                # 모듈간 공종들
                elif pre_workpackage =="Z" and suc_workpackage == "Z":
                    pre_work = get_work(pre, works_of_section)[0]
                    suc_work = get_work(suc, works_of_section)[0]
                    dep_list.append([pre_work["work_id"], suc_work["work_id"]])
            ctn += 1
    return dep_list


def get_last_tasktype(dir_txt):
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel["last_tasktype_raw"]    
    last_tasktype_list = []
    for row in excel_ws.iter_rows():                           
        last_tasktype_list.append(row[0].value)       

    return last_tasktype_list


def generate_json(dir_txt, generated_dir):
    labor_list = get_labor_type_list(dir_txt)
    print("generate labor_list")
    task_type_list = get_tasktype_list(dir_txt)
    print("generate task_type_list")
    work_list = get_work_list(dir_txt)
    print("generate work_list")
    space_list = get_space_list(dir_txt)
    print("generate dep_list")
    dep_list = get_dependency_list(dir_txt)
    last_tasktype_id_list = get_last_tasktype(dir_txt)
    data = {
        "section": get_section_list(dir_txt),
        "workpackage": get_workpackage_list(dir_txt),
        "last_tasktype_id": last_tasktype_id_list,
        "labor_type": labor_list,
        "task_type": task_type_list,
        "work": work_list,
        "space": space_list,
        "dependency": dep_list
    }
    with open(generated_dir, 'w', encoding="utf-8") as make_file:
        json.dump(data, make_file, ensure_ascii=False, indent="\t")
    print("json_file is generated at " + generated_dir)
    return
    
