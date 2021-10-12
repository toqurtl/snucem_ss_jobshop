import collections
from ortools.sat.python import cp_model
import openpyxl

dir_txt = "data/input.xlsx"
# read_space_raw
def get_space_list(dir_txt):
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel['space_raw']
    ctn = 0
    space_list = []
    for row in excel_ws.iter_rows():
        if ctn > 0:
            module_type, zone, num_section, quantity = \
                row[0].value, row[1].value, int(row[2].value), int(row[3].value)
            
            space_list.append(
                {
                    "space_id": zone,
                    "module_type": "module",
                    "upper_space": None,
                    "detail": 0,
                    "is_module": True,
                    "quantity": 1
                }
            )
            for idx in range(1, num_section + 1):
                space_list.append(
                    {
                        "space_id": zone+"_"+str(idx),
                        "module_type": module_type,
                        "upper_space": zone,
                        "detail": idx,
                        "is_module": False,
                        "quantity": round(quantity / num_section)
                    }
                )
        ctn += 1
    return space_list

def get_tasktype_list(dir_txt):
    excel = openpyxl.load_workbook(dir_txt)
    excel_ws = excel['tasktype_raw']
    ctn = 0
    tasktype_list = []    
    for row in excel_ws.iter_rows():        
        if ctn > 0:
            work_id, workpackage_id, workpackage_name, detail_id, detail_name, is_module, productivity \
                = row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, bool(row[5].value), float(row[6].value)
            tasktype_list.append({
                "task_id": work_id,
                "workpackage_id": workpackage_id,
                "workpackage_name": workpackage_name,
                "detail_id": detail_id,
                "task_name": detail_name,
                "is_module": is_module,
                "productivity": productivity
            })
        ctn += 1
    return tasktype_list

def get_low_space_list(space_id, space_list):
    low_space_list = []
    for space in space_list:
        if space["upper_space"] == space_id:
            low_space_list.append(space["space_id"])
    return low_space_list


# print(get_space_list(dir_txt))
space_list = get_space_list(dir_txt)
tasktype_list = get_tasktype_list(dir_txt)
work_list = []
for space in space_list:    
    space_id = space["space_id"]    
    for tasktype in tasktype_list:
        if space["is_module"]:
            if tasktype["is_module"]:
                work_id = space_id + "_" + tasktype["task_id"]                
                work_list.append({
                    "work_id": work_id,
                    "task_id": tasktype["task_id"],
                    "quantity": space["quantity"],                    
                    "space_id_list": get_low_space_list(space_id, space_list)
                })
                
        else:
            if not tasktype["is_module"]:
                work_id = space_id + "_" + tasktype["task_id"]
                work_list.append({
                    "work_id": work_id,
                    "task_id": tasktype["task_id"],
                    "quantity": space["quantity"],
                    "space_id_list": [space_id]
                })
for work in work_list:
    print(work)        
        